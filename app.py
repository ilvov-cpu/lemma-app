import re
from collections import defaultdict
from io import BytesIO

import pandas as pd
import pymorphy3
import streamlit as st
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Агрегация слов по леммам", layout="wide")

DEFAULT_STOP_WORDS = {
    "в", "на", "и", "с", "по", "для", "из", "от", "до", "у", "о", "об", "под", "над"
}


@st.cache_resource
def get_morph_analyzer():
    return pymorphy3.MorphAnalyzer(lang="ru")


MORPH = get_morph_analyzer()


def extract_words(text: str) -> list[str]:
    if pd.isna(text):
        return []
    text = str(text).lower()
    return re.findall(r"[а-яёa-z0-9]+", text)


@st.cache_data
def normalize_word(word: str) -> str:
    parsed = MORPH.parse(word)
    if not parsed:
        return word
    return parsed[0].normal_form


def aggregate_by_lemma_from_df(
    df: pd.DataFrame,
    stop_words: set[str],
    count_once_per_key: bool = True,
) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.str.strip()

    required_columns = {"Ключ", "Расход", "Доход"}
    missing = required_columns - set(df.columns)
    if missing:
        raise ValueError(f"В файле отсутствуют колонки: {', '.join(sorted(missing))}")

    df["Расход"] = pd.to_numeric(df["Расход"], errors="coerce").fillna(0)
    df["Доход"] = pd.to_numeric(df["Доход"], errors="coerce").fillna(0)

    stats = defaultdict(lambda: {
        "Расход": 0.0,
        "Доход": 0.0,
        "Формы": set(),
        "Количество ключей": 0,
    })

    for _, row in df.iterrows():
        words = extract_words(row["Ключ"])

        if count_once_per_key:
            lemmas_in_key = {}
            for word in words:
                lemma = normalize_word(word)
                if lemma not in stop_words:
                    lemmas_in_key.setdefault(lemma, set()).add(word)

            for lemma, forms in lemmas_in_key.items():
                stats[lemma]["Расход"] += row["Расход"]
                stats[lemma]["Доход"] += row["Доход"]
                stats[lemma]["Количество ключей"] += 1
                stats[lemma]["Формы"].update(forms)
        else:
            for word in words:
                lemma = normalize_word(word)
                if lemma in stop_words:
                    continue
                stats[lemma]["Расход"] += row["Расход"]
                stats[lemma]["Доход"] += row["Доход"]
                stats[lemma]["Количество ключей"] += 1
                stats[lemma]["Формы"].add(word)

    result = pd.DataFrame([
        {
            "Слово": lemma,
            "Расход": values["Расход"],
            "Доход": values["Доход"],
            "ДРР, %": round((values["Расход"] / values["Доход"] * 100), 2) if values["Доход"] else None,
            "Количество ключей": values["Количество ключей"],
            "Найденные формы": ", ".join(sorted(values["Формы"])),
        }
        for lemma, values in stats.items()
    ])

    if result.empty:
        return result

    result = result.sort_values(by="Доход", ascending=False).reset_index(drop=True)
    return result


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df = df.copy()
        export_df.to_excel(writer, index=False, sheet_name="Результат")

        ws = writer.sheets["Результат"]

        for idx, col_name in enumerate(export_df.columns, start=1):
            col_letter = get_column_letter(idx)
            header_len = len(str(col_name))

            if export_df.empty:
                max_len = header_len
            else:
                series = export_df.iloc[:, idx - 1]
                cell_len = series.map(lambda x: len(str(x)) if pd.notna(x) else 0).max()
                max_len = max(header_len, int(cell_len))

            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        for row in range(2, len(export_df) + 2):
            ws[f"B{row}"].number_format = "#,##0.00"
            ws[f"C{row}"].number_format = "#,##0.00"
            ws[f"D{row}"].number_format = "0.00"

    output.seek(0)
    return output.getvalue()


st.title("Агрегация слов по леммам")
st.write(
    "Загрузи Excel-файл с колонками **Ключ**, **Расход**, **Доход**. "
    "Приложение объединит словоформы: например, **Москва** и **Москве**."
)

with st.sidebar:
    st.header("Настройки")
    remove_stop_words = st.checkbox("Убирать стоп-слова", value=True)
    count_once_per_key = st.checkbox("Считать слово один раз на ключ", value=True)
    custom_stop_words = st.text_area(
        "Дополнительные стоп-слова",
        value="",
        placeholder="Например: купить, заказать, цена",
        help="Укажи слова через запятую. Они будут исключены из результата.",
    )

uploaded_file = st.file_uploader("Выбери Excel-файл", type=["xlsx"])

if uploaded_file is not None:
    try:
        source_df = pd.read_excel(uploaded_file)

        st.subheader("Предпросмотр исходных данных")
        st.dataframe(source_df.head(10), use_container_width=True)

        if st.button("Обработать файл", type="primary"):
            stop_words = set(DEFAULT_STOP_WORDS) if remove_stop_words else set()
            if custom_stop_words.strip():
                stop_words.update(
                    {w.strip().lower() for w in custom_stop_words.split(",") if w.strip()}
                )

            result_df = aggregate_by_lemma_from_df(
                source_df,
                stop_words=stop_words,
                count_once_per_key=count_once_per_key,
            )

            if result_df.empty:
                st.warning("Результат пустой. Проверь входные данные или настройки стоп-слов.")
            else:
                st.subheader("Результат")
                st.dataframe(result_df, use_container_width=True)

                excel_bytes = dataframe_to_excel_bytes(result_df)

                st.download_button(
                    label="Скачать результат в Excel",
                    data=excel_bytes,
                    file_name="output_lemmas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    except Exception as e:
        st.error(f"Ошибка: {e}")
else:
    st.info("Сначала загрузи файл .xlsx")
